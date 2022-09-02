from typing import List
import openpyxl
import openpyxl.styles
from openpyxl.utils.cell import get_column_letter

from pkg.trackRepo import TrackRepo

center_align = openpyxl.styles.Alignment(
    horizontal='center', vertical='center', wrap_text=True)

START_CHECK_COL = 3
START_CHECK_ROW = 4


class HandlerExcel():
    def __init__(self):
        self.fileName = 'GSW_Tracking.xlsx'
        self.workbook = openpyxl.Workbook()

    def ExportData(self, trackRepo: TrackRepo, clone_list: List[str]):
        sheet = self.workbook.active
        # Title
        sheet.merge_cells('A1:AD1')
        sheet['A1'] = 'TRACKING CTA'
        title_cell = sheet.cell(row=1, column=1)
        title_cell.alignment = center_align
        title_cell.font = openpyxl.styles.Font(
            bold=True, color='00FF0000', size=20)

        # Add Header from row 3 - style all yellow
        self.all_CTA_date = trackRepo.GetAllMatchTime()
        self.total_CTA = len(self.all_CTA_date)
        map_date_col = {self.all_CTA_date[i]: i + 3
                        for i in range(self.total_CTA)}

        sheet['A3'] = 'STT'
        sheet['B3'] = 'Name'
        sheet.column_dimensions['B'].width = 25
        sheet.cell(row=3, column=START_CHECK_COL +
                   self.total_CTA).value = 'Total'

        for i in range(self.total_CTA):
            sheet.cell(row=3, column=START_CHECK_COL +
                       i).value = self.all_CTA_date[i].replace(' ', '\n')
            sheet.column_dimensions[get_column_letter(i + 3)].width = 14

        header_font = openpyxl.styles.Font(name="Arial", size=12, bold=True)
        for cell in sheet["3:3"]:
            cell.font = header_font
            cell.alignment = center_align
            # Fill yellow
            cell.fill = openpyxl.styles.PatternFill(
                start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")

        # ID - Player name cols
        self.players_name = trackRepo.GetAllPlayersName()
        num_players = len(self.players_name)
        for i in range(num_players):
            sheet.cell(row=i+START_CHECK_ROW, column=1).value = i + 1
            sheet.cell(row=i+START_CHECK_ROW,
                       column=1).alignment = center_align
        for i in range(num_players):
            sheet.cell(row=i+START_CHECK_ROW,
                       column=2).value = self.players_name[i]

        # Checkmark attendance by row
        for idPlayer, player in enumerate(self.players_name):
            # Clone acc
            if player in clone_list:
                for cell in sheet["{row}:{row}".format(row = START_CHECK_ROW + idPlayer)]:
                    # Fill green
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='00007F00', end_color='00007F00', fill_type="solid")
                continue

            allDate = trackRepo.GetAllDateOfPlayer(player)
            for date in allDate:
                check_cell = sheet.cell(
                    row=START_CHECK_ROW + idPlayer, column=map_date_col[date])
                check_cell.value = 'X'
                check_cell.alignment = center_align

            # Summarize attend over total CTA
            total_cell = sheet.cell(row=START_CHECK_ROW + idPlayer,
                                    column=START_CHECK_COL + self.total_CTA)

            percent_attend = (len(allDate) / self.total_CTA) * 100
            total_cell.value = "{} / {} ({:.2f}%)".format(len(allDate),
                                                        self.total_CTA, percent_attend)
            total_cell.alignment = center_align
            total_cell.font = openpyxl.styles.Font(size=14, bold=True)
            total_cell.fill = openpyxl.styles.PatternFill(
                start_color="00CC99FF", end_color="00CC99FF", fill_type="solid")

        self.workbook.save(self.fileName)
